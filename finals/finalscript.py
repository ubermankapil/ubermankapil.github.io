import openpyxl, smtplib, sys
reload(sys)
sys.setdefaultencoding('utf-8')

SUBJECT = 'Summer Internship Under You'

gmail_sender = 'kapilinfuse@gmail.com'
gmail_passwd = 'kapil2011'
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(gmail_sender, gmail_passwd)

wb = openpyxl.load_workbook('hero.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

for r in range(2, sheet.get_highest_row()+1):
	name = sheet.cell(row=r, column=1).value
	interest = sheet.cell(row=r, column=2).value
	email = sheet.cell(row=r, column=3).value

	body = "Dear Prof. {0},\n \nI am Kapil Agrawal, an undergraduate studying Bachelors in Electrical and Electronics Engineering and Masters in Economics at Birla Institute of Technology and Science, Pilani (BITS Pilani, India).\n \nI was looking into your current research interests and would be keenly interested in working with you from May-July, 2016 in {1}. An internship opportunity at this stage shall be very helpful for me in building a greater insight to the subjects of my study. \n\nI assure you of my strong academic and technical aptitude. I have experience and have done many projects in Computer Programming, Web development and Android App development. I also hold a very good academic record. I was ranked among the top 0.4%  engineering students in India (5717 out of about 1.4 million students in IIT-JEE 2014). Also, I am a recipient of National Talent Search Scholarship awarded by Government of India to top 1000 students out of 0.3 million students after clearing three rounds of examinations and interviews. \n\nMy resume can be checked out here - http://ubermankapil.github.io. If you feel I require any additional skills or do any pre-requisite readings, I will gladly do them in preparation. If you require any further information, I would be more than willing to furnish the same. \n\nThank you for your time. Let me know if any potential opportunities are available in your group. Looking forward to hearing from you. \n\nRegards, \nKapil Agrawal, \nBits Pilani.".format(name, interest)
	print ('Sending email to %s...' % email)

	BODY = '\r\n'.join([
		'Subject: %s' % SUBJECT,
		'',
		body
		])
	try:
		smtpObj.sendmail(gmail_sender, email, BODY)
		print 'email sent'
	except:
		print 'error sending email'

smtpObj.quit()
