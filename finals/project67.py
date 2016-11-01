import openpyxl, sys, xlsxwriter
reload(sys)
sys.setdefaultencoding('utf-8')


wb = openpyxl.load_workbook('67.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

workbook = xlsxwriter.Workbook('BanchoBB.xlsx')
worksheet = workbook.add_worksheet()

array1 = []
 

for r in range(2, sheet.get_highest_row()+1):
	array1.append(str(sheet.cell(row=r, column=1).value))

print array1

for r in range(2, 65537):
	if str(sheet.cell(row=r, column=3).value) in array1:
		worksheet.write(r,0,str("1"))
	else:
		worksheet.write(r,0,str("0"))

workbook.close()
print ("Values alloted")
		

	
